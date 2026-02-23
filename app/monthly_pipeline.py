import pandas as pd
from difflib import get_close_matches
import unicodedata
import re
import io
import threading
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

from app.read_excel_file import process_excel
from app.read_pdf_file import process_pdf

THICK_BORDER = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="medium"),
    bottom=Side(style="medium"),
)

def auto_adjust_column_width(ws, padding=4):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + padding

def process_final(procare_file, dhs_file, auth_file):

    print("Excel ve PDF iÅŸlemleri baÅŸlatÄ±lÄ±yor (threading)...")

    excel_result = {}
    pdf_result = {}

    def run_excel():
        excel_result["data"] = process_excel(procare_file)

    def run_pdf():
        pdf_result["data"] = process_pdf(dhs_file)

    excel_thread = threading.Thread(target=run_excel)
    pdf_thread = threading.Thread(target=run_pdf)

    excel_thread.start()
    pdf_thread.start()

    excel_thread.join()
    pdf_thread.join()

    print("âœ” Excel ve PDF iÅŸlemleri tamamlandÄ±.\n")

    first = excel_result["data"]
    second = pdf_result["data"]

#   ðŸ”§ Clean column names
    second.columns = second.columns.str.strip().str.upper()

    # -----------------------------
    # NAME CLEANING FUNCTIONS
    # -----------------------------

    def clean_name(name):
        name = unicodedata.normalize('NFKD', name)
        name = name.upper()
        name = re.sub(r'[^A-Z ]', '', name)
        name = re.sub(r'\s+', ' ', name)
        return name.strip()

    def name_matches(second_fullname, first_fullname):
        s = clean_name(second_fullname).split()
        f = clean_name(first_fullname).split()
        return s == f

    # -----------------------------
    # DATE FORMAT NORMALIZATION
    # -----------------------------

    first["Attdate"] = pd.to_datetime(first["Attdate"], errors="coerce").dt.strftime("%m/%d/%Y")
    second["SWIPE DATE"] = pd.to_datetime(second["SWIPE DATE"], errors="coerce").dt.strftime("%m/%d/%Y")

    second["NOTE"] = ""
    second["COPAY EXTRA"] = None
    second["AMOUNT EXTRA"] = None

    # -----------------------------
    # MATCHING LOGIC
    # -----------------------------

    final_rows = []

    for _, row in second.iterrows():
        full_name = str(row["FULL NAME"]).strip().upper()
        swipe_date = str(row["SWIPE DATE"]).strip()

        possible_matches = [
            name for name in first["Full Name"].tolist()
            if name_matches(full_name, name)
        ]

        copay_val = float(row.get("COPAY AP", 0) or 0)
        amount_val = float(row.get("AMOUNT PAID", 0) or 0)

        if possible_matches:
            matched_rows = first[first["Full Name"].isin(possible_matches)]

            if swipe_date in matched_rows["Attdate"].values:
                row["NOTE"] = "DHS PAID"
                final_rows.append(row)
            else:
                if amount_val == 14.00:
                    row["NOTE"] = "NON TRADITIONAL"
                    row["COPAY EXTRA"] = copay_val
                    row["AMOUNT EXTRA"] = amount_val
                    final_rows.append(row)
                    continue

                if copay_val == 0 and amount_val == 0:
                    continue

                if copay_val > 0 and amount_val > 0:
                    row["NOTE"] = "EXTRA COPAY & EXTRA DHS"
                elif amount_val > 0:
                    row["NOTE"] = "EXTRA DHS"
                elif copay_val > 0:
                    row["NOTE"] = "EXTRA COPAY"
                else:
                    row["NOTE"] = "EXTRA"

                row["COPAY EXTRA"] = copay_val
                row["AMOUNT EXTRA"] = amount_val
                final_rows.append(row)
        else:
            if copay_val == 0 and amount_val == 0:
                continue

            if copay_val > 0 and amount_val > 0:
                row["NOTE"] = "EXTRA COPAY & EXTRA DHS"
            elif amount_val > 0:
                row["NOTE"] = "EXTRA DHS"
            elif copay_val > 0:
                row["NOTE"] = "EXTRA COPAY"
            else:
                row["NOTE"] = "EXTRA"

            row["COPAY EXTRA"] = copay_val
            row["AMOUNT EXTRA"] = amount_val
            final_rows.append(row)

    second = pd.DataFrame(final_rows)

# -----------------------------
# ADDITIONAL LOOP: SELF PAID / NOT PAID
# -----------------------------
    extra_rows = []

    for _, row in first.iterrows():
        full_first = str(row["Full Name"]).strip().upper()
        date_first = str(row["Attdate"]).strip()

        possible_matches = [
            name for name in second["FULL NAME"].tolist()
            if name_matches(full_first, name)
        ]

        if not possible_matches:
            # CASE 1: SELF PAID
            extra_rows.append({
                "FULL NAME": full_first,
                "CASE/PERSON": "NO CASE#",
                "SWIPE DATE": date_first,
                "COPAY AP": "",
                "AMOUNT PAID": "",
                "NOTE": "SELF PAID",
                "COPAY EXTRA": "",
                "AMOUNT EXTRA": ""
            })
            continue

        matched_rows = second[second["FULL NAME"].isin(possible_matches)]

        # CASE 2: exists by name but NOT on that date â†’ NOT PAID
        if date_first not in matched_rows["SWIPE DATE"].astype(str).values:
            case_person = matched_rows.iloc[0].get("CASE/PERSON", "NO CASE#")
            if pd.isna(case_person) or case_person == "":
                case_person = "NO CASE#"

            extra_rows.append({
                "FULL NAME": full_first,
                "CASE/PERSON": case_person,
                "SWIPE DATE": date_first,
                "COPAY AP": 0.0,
                "AMOUNT PAID": 0.0,
                "NOTE": "NOT PAID",
                "COPAY EXTRA": "",
                "AMOUNT EXTRA": ""
            })

    if extra_rows:
        extra_df = pd.DataFrame(extra_rows)
        second = pd.concat([second, extra_df], ignore_index=True)

    # -----------------------------
    # NAME NORMALIZATION
    # -----------------------------
    unique_names = second["FULL NAME"].unique()
    name_map = {}

    for name in unique_names:
        match = get_close_matches(name, list(name_map.keys()), cutoff=0.85)
        if match:
            name_map[name] = name_map[match[0]]
        else:
            name_map[name] = name

    second["FULL NAME"] = second["FULL NAME"].map(name_map)
    second["FULL NAME"] = second["FULL NAME"].str.strip()
    second["FULL NAME"] = second["FULL NAME"].str.replace(r"\s+", " ", regex=True)

    # -----------------------------
    # SORTING + DUPLICATE CONTROL
    # -----------------------------
    second["SWIPE DATE"] = pd.to_datetime(second["SWIPE DATE"], errors="coerce")
    second = second.sort_values(by=["FULL NAME", "SWIPE DATE"], ascending=[True, True])
    second["SWIPE DATE"] = second["SWIPE DATE"].dt.strftime("%m/%d/%Y")

    duplicates = second.duplicated(subset=["FULL NAME", "SWIPE DATE"], keep=False)

    for (name, date), group in second[duplicates].groupby(["FULL NAME", "SWIPE DATE"]):
        zero_rows = group[(group["AMOUNT PAID"] == 0) & (group["COPAY AP"] == 0)]
        if not zero_rows.empty:
            second = second.drop(zero_rows.index)

    # -----------------------------
    # DHS PAID -> NOT PAID dÃ¼zeltme
    # -----------------------------
    for idx, row in second.iterrows():
        if str(row.get("NOTE", "")).upper() == "DHS PAID" and \
        float(row.get("AMOUNT PAID", 0) or 0) == 0 and \
        float(row.get("COPAY AP", 0) or 0) == 0:
            second.at[idx, "NOTE"] = "NOT PAID"


    # -----------------------------
    # AUTHORIZATION CHECK
    # -----------------------------

    auth_list = pd.read_excel(auth_file, dtype=str)
    auth_list.columns = auth_list.columns.str.strip().str.upper()

    for idx, row in second.iterrows():
        if str(row.get("NOTE", "")).upper() == "SELF PAID":
            student_name = str(row["FULL NAME"]).strip().upper()

            auth_match = auth_list[
                auth_list["CHILD NAME"].apply(lambda x: name_matches(student_name, str(x)))
            ]

            if not auth_match.empty:
                second.at[idx, "NOTE"] = "NOT PAID"
                second.at[idx, "COPAY AP"] = 0
                second.at[idx, "AMOUNT PAID"] = 0

                case_number = str(auth_match.iloc[0]["CASE #"]).strip()
                person_number = str(auth_match.iloc[0]["PERSON"]).strip()

                second.at[idx, "CASE/PERSON"] = f"{case_number}/{person_number}"

    for i, row in second.iterrows():
        if row["AMOUNT PAID"] == 14:
            second.at[i, "NOTE"] = "NON TRADITIONAL"
            second.at[i, "COPAY EXTRA"] = 0
            second.at[i, "AMOUNT EXTRA"] = 14


    # -----------------------------
    # MEMORY EXCEL EXPORT
    # -----------------------------

    output = io.BytesIO()
    second.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    note_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "NOTE":
            note_col = idx
            break

    red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    for row in ws.iter_rows(min_row=2):
        note_cell = row[note_col - 1]
        note_value = str(note_cell.value).strip().upper()

        if note_value == "NOT PAID":
            note_cell.fill = red
        elif note_value == "NON TRADITIONAL":
            note_cell.fill = yellow
        elif note_value == "SELF PAID":
            continue  # No color (aynen senin mantÄ±ÄŸÄ±n)
        elif note_value in ["EXTRA DHS", "EXTRA COPAY", "EXTRA COPAY & EXTRA DHS"]:
            note_cell.fill = yellow
        else:
            note_cell.fill = green

    # ===== APPLY THICK BORDERS =====
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column
    ):
        for cell in row:
            cell.border = THICK_BORDER

    # ===== AUTO COLUMN WIDTH =====
    auto_adjust_column_width(ws, padding=4)

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    print("ðŸŽ¨ Renkler uygulandÄ±.")
    print("âœ… Final rapor hazÄ±r.")

    return final_output